from openpyxl import load_workbook
import openpyxl
import pprint
from collections import Counter


lis = {}
var = []
total = []
variants = []
boxes = {}


def Calc1():
    wbopen = load_workbook('./Датасет.xlsm')
    sheet = wbopen['Таблица']
    
    for i in range(2, sheet.max_row + 1):
        form = sheet['B'+str(i)].value
        color = sheet['C'+str(i)].value
        size = sheet['D'+str(i)].value
        lumen = sheet['E'+str(i)].value
        
        if size == "Большой":
            size = "3Большой"
        elif size == "Средний":
            size = "2Средний"
        elif size == "Малый":
            size = "1Малый"
        var.append(size+form+color+lumen)
        print(str((i-1)/10)+'%')



Calc1()
print("----------------------------------------------------------------")
print("данные загружены")
print("----------------------------------------------------------------")


while True:
    end = []
    newvar = set(var)

    number = int(input('Введите количество боксов -> '))

    for el in newvar:
        variants.append(el)

    var.sort()

    total = {i:var.count(i) for i in var}


    for e in range(1, number +1):
        boxes[e] = list()

    count = 1
    #pprint.pprint(boxes)

    for i in range(len(var)):
        boxes[count].append(var[i])
        count = count + 1
        if count > number:
            count = 1
        

    #pprint.pprint(boxes)
    #print("-------------------------------------------------------------")
            
    #pprint.pprint(slov)
    #sortedDict = sorted(slov.items(), key=lambda x: x[1])


    for i in range(1, number + 1):
        countbox = 0
        slov = list(boxes[i])
        array_d = {}.fromkeys(slov, 0)
        for a in slov:
            array_d[a] += 1
            countbox = countbox + 1
        end.append(countbox)
        print('Бокс №', i)
        pprint.pprint(array_d)
        
        print('')
        print('Количество алмазов в боксе №', i, " - ", countbox, "шт.")
        print("----------------------------------------------------------------")

    for i in range(len(end)):
        print('Бокс №', i + 1, ' - ', end[i], "шт.")

    print(" ")
    print(" ")
    

